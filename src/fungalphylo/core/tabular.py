from __future__ import annotations

import csv
import gzip
from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path

import openpyxl


@dataclass(frozen=True)
class Table:
    path: Path
    kind: str  # "delimited" or "excel"
    delimiter: str | None
    fieldnames: list[str]


def _open_text(path: Path):
    path = path.expanduser().resolve()
    if path.suffix == ".gz":
        return gzip.open(path, "rt", encoding="utf-8", newline="")
    return path.open("r", encoding="utf-8", newline="")


def sniff_delimiter(sample: str) -> str:
    if "\t" in sample and sample.count("\t") >= sample.count(","):
        return "\t"
    try:
        return csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";", "|"]).delimiter
    except Exception:
        return "\t"


def _read_excel(path: Path, sheet: str | None = None) -> tuple[Table, Iterator[dict[str, str]]]:
    """
    Read first sheet (or named sheet) of an Excel workbook.
    Includes hyperlink targets as extra columns: <col>__link
    """
    path = path.expanduser().resolve()
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]

    # Header row
    header_cells = list(ws.iter_rows(min_row=1, max_row=1))[0]
    headers = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
    if not any(headers):
        raise ValueError(f"{path} has an empty header row.")
    fieldnames = headers.copy()

    # Add link columns only when present in any row; simplest is to always include them
    link_fieldnames = [f"{h}__link" for h in headers if h]
    fieldnames_extended = [h for h in fieldnames if h] + link_fieldnames

    meta = Table(path=path, kind="excel", delimiter=None, fieldnames=fieldnames_extended)

    def row_iter() -> Iterator[dict[str, str]]:
        for row in ws.iter_rows(min_row=2):
            out: dict[str, str] = {}
            empty = True
            for h, cell in zip(headers, row):
                if not h:
                    continue
                val = cell.value
                sval = "" if val is None else str(val).strip()
                if sval:
                    empty = False
                out[h] = sval

                link = cell.hyperlink.target if cell.hyperlink else ""
                if link:
                    out[f"{h}__link"] = str(link).strip()
                    empty = False
                else:
                    out[f"{h}__link"] = ""
            if empty:
                continue
            yield out

    return meta, row_iter()


def read_table(path: Path, delimiter: str | None = None) -> tuple[Table, Iterator[dict[str, str]]]:
    """
    Return (Table metadata, iterator over rows as dict[str,str]).
    Supports:
      - TSV/CSV (optionally .gz)
      - Excel .xlsx/.xlsm (preserves hyperlinks as <col>__link)
    """
    path = path.expanduser().resolve()
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _read_excel(path)

    # Delimited text
    f = _open_text(path)
    sample = f.read(4096)
    f.seek(0)

    delim = delimiter or sniff_delimiter(sample)
    reader = csv.DictReader(f, delimiter=delim)
    if reader.fieldnames is None:
        f.close()
        raise ValueError(f"{path} appears to have no header row.")

    meta = Table(path=path, kind="delimited", delimiter=delim, fieldnames=list(reader.fieldnames))

    def row_iter() -> Iterator[dict[str, str]]:
        try:
            for row in reader:
                yield {k.strip(): (v.strip() if isinstance(v, str) else "") for k, v in row.items() if k is not None}
        finally:
            f.close()

    return meta, row_iter()